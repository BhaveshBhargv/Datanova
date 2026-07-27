"""Tests for Phase 10 reporting & export."""
import io

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

LIST = "/api/datasets"


def _upload_df(client, headers, df: pd.DataFrame) -> dict:
    buf = df.to_csv(index=False).encode()
    return client.post(
        f"{LIST}/upload", files={"file": ("d.csv", buf, "text/csv")}, headers=headers
    ).json()


def _df(n: int = 80) -> pd.DataFrame:
    rng = np.random.default_rng(0)
    f1 = rng.normal(size=n)
    return pd.DataFrame(
        {"f1": f1, "f2": rng.normal(size=n), "label": (f1 > 0).astype(int)}
    )


@pytest.fixture()
def force_fallback(monkeypatch):
    monkeypatch.setattr("app.services.narrate.llm.generate", lambda *a, **k: None)


def test_report_json_has_sections(client, auth_headers, force_fallback):
    headers = auth_headers()
    ds = _upload_df(client, headers, _df())
    r = client.get(f"{LIST}/{ds['id']}/report", headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    for key in ("dataset", "profile", "eda", "insights", "summary", "preview"):
        assert key in body
    assert body["dataset"]["name"]
    assert body["summary"]["overview_source"] == "fallback"


def test_report_includes_model_when_trained(client, auth_headers, force_fallback):
    headers = auth_headers()
    ds = _upload_df(client, headers, _df())
    exp = client.post(
        f"{LIST}/{ds['id']}/experiments", json={"target": "label"}, headers=headers
    ).json()
    assert exp["status"] == "completed"
    body = client.get(f"{LIST}/{ds['id']}/report", headers=headers).json()
    assert body["experiment"] is not None
    assert body["experiment"]["best_model_name"]
    assert body["importance"] is not None  # SHAP importances present


def test_pdf_download(client, auth_headers, force_fallback):
    headers = auth_headers()
    ds = _upload_df(client, headers, _df())
    r = client.get(f"{LIST}/{ds['id']}/report/pdf", headers=headers)
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:4] == b"%PDF"
    assert len(r.content) > 1000
    assert "attachment" in r.headers.get("content-disposition", "")


def test_excel_download_is_valid_workbook(client, auth_headers, force_fallback):
    headers = auth_headers()
    ds = _upload_df(client, headers, _df())
    r = client.get(f"{LIST}/{ds['id']}/report/excel", headers=headers)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    assert {"Overview", "Column Profile", "Insights", "Data Preview"} <= set(wb.sheetnames)


def test_report_ownership_enforced(client, auth_headers):
    alice = auth_headers("alice@example.com")
    bob = auth_headers("bob@example.com")
    ds = _upload_df(client, alice, _df())
    assert client.get(f"{LIST}/{ds['id']}/report", headers=bob).status_code == 404
    assert client.get(f"{LIST}/{ds['id']}/report/pdf", headers=bob).status_code == 404
    assert client.get(f"{LIST}/{ds['id']}/report/excel", headers=bob).status_code == 404
