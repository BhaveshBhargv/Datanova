"""Render an assembled report dict to a multi-sheet Excel workbook (openpyxl)."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

_HEADER = Font(bold=True)
_METRIC_ORDER = {
    "classification": ["accuracy", "precision", "recall", "f1", "roc_auc"],
    "regression": ["r2", "rmse", "mae"],
}


def _header(ws, row: list):
    ws.append(row)
    for cell in ws[ws.max_row]:
        cell.font = _HEADER


def build_excel(report: dict) -> bytes:
    wb = Workbook()
    ds = report["dataset"]
    prof = report["profile"]

    # --- Overview ---
    ws = wb.active
    ws.title = "Overview"
    _header(ws, ["Metric", "Value"])
    for label, value in [
        ("Dataset", ds["name"]),
        ("Rows", ds["n_rows"]),
        ("Columns", ds["n_columns"]),
        ("Source", ds["source_type"]),
        ("Quality score", prof["quality_score"]),
        ("Missing cells", prof["missing_cells"]),
        ("Missing %", prof["missing_pct"]),
        ("Duplicate rows", prof["duplicate_rows"]),
    ]:
        ws.append([label, value])
    ws.append([])
    ws.append(["Executive summary"])
    ws.append([report["summary"]["overview"]])
    ws.append([report["summary"]["insights"]])

    # --- Column Profile ---
    ws = wb.create_sheet("Column Profile")
    _header(ws, ["Column", "Type", "Count", "Missing", "Missing %", "Unique", "Outliers"])
    for c in prof["columns"]:
        ws.append(
            [c["name"], c["dtype"], c["count"], c["missing"], c["missing_pct"], c["unique"], c.get("outliers")]
        )

    # --- Correlations ---
    corr = report["eda"]["correlations"]
    if corr.get("columns"):
        ws = wb.create_sheet("Correlations")
        _header(ws, ["", *corr["columns"]])
        for name, row in zip(corr["columns"], corr["matrix"]):
            ws.append([name, *row])

    # --- Insights ---
    ws = wb.create_sheet("Insights")
    _header(ws, ["Severity", "Category", "Title", "Detail", "Recommendation"])
    for ins in report["insights"]:
        ws.append([ins["severity"], ins["category"], ins["title"], ins["detail"], ins.get("recommendation")])

    # --- Model Metrics ---
    exp = report.get("experiment")
    if exp:
        ws = wb.create_sheet("Model Metrics")
        ws.append(["Target", exp["target"], "Problem", exp["problem_type"]])
        ws.append(["Best model", exp["best_model_name"]])
        ws.append([])
        metric_keys = [
            k
            for k in _METRIC_ORDER.get(exp["problem_type"], [])
            if any(k in r["metrics"] for r in exp["results"])
        ]
        _header(ws, ["Model", *metric_keys])
        for r in exp["results"]:
            ws.append([r["model"], *[r["metrics"].get(k) for k in metric_keys]])
        if report.get("importance"):
            ws.append([])
            _header(ws, ["Feature", "Importance (mean |SHAP|)"])
            for item in report["importance"]:
                ws.append([item["feature"], item["importance"]])

    # --- Data Preview ---
    ws = wb.create_sheet("Data Preview")
    preview = report["preview"]
    _header(ws, preview["columns"])
    for row in preview["rows"]:
        ws.append([row.get(c) for c in preview["columns"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
